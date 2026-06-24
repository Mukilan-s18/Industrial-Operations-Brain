"""
Generate synthetic demo documents for the Industrial Operations Brain demo.
Run: python demo_docs/generate_demo_docs.py
"""

import os
from pathlib import Path

OUTPUT_DIR = Path(__file__).parent

# ── Requires PyMuPDF ──────────────────────────────────────────────────────────
import fitz


def _new_pdf(filename: str, pages_content: list[str]) -> Path:
    doc = fitz.open()
    for content in pages_content:
        page = doc.new_page(width=595, height=842)  # A4
        page.insert_text((72, 72), content, fontsize=11)
    out = OUTPUT_DIR / filename
    doc.save(str(out))
    doc.close()
    print(f"  Created: {out.name}")
    return out


def create_sop_rev3():
    """SOP Rev 3 — outdated, will trigger version conflict."""
    _new_pdf("sop_pump_p101_rev3.pdf", [
        """STANDARD OPERATING PROCEDURE
Equipment: Centrifugal Pump P-101
Revision 3
Date: 10/06/2023
Approved by: Engineering Manager

SCOPE
This procedure covers routine maintenance and inspection of pump P-101.

SAFETY PRECAUTIONS
- Isolate all energy sources before beginning work
- Wear PPE at all times (gloves, safety glasses, hard hat)
- Verify zero-energy state before opening pump casing

PROCEDURE
Step 1: Shut down pump P-101 using control panel HMI.
Step 2: Close suction valve SV-101 and discharge valve DV-101.
Step 3: Drain pump casing via drain valve.
Step 4: Inspect impeller for wear or damage.
Step 5: Check mechanical seal for leaks.
Step 6: Torque bolts to 50 Nm per specification.
Step 7: Reconnect and perform leak test at 80% operating pressure.

SIGN-OFF
Technician: _______________  Date: _______________
Supervisor: _______________  Date: _______________
""",
        """REVISION HISTORY
Rev 3 | 10/06/2023 | Updated safety section
Rev 2 | 15/03/2022 | Added drain step
Rev 1 | 01/01/2021 | Initial release

EQUIPMENT TAGS REFERENCED
P-101, SV-101, DV-101, HV-204
"""
    ])


def create_sop_rev4():
    """SOP Rev 4 — current version with updated torque spec."""
    _new_pdf("sop_pump_p101_rev4.pdf", [
        """STANDARD OPERATING PROCEDURE
Equipment: Centrifugal Pump P-101
Revision 4
Date: 15/01/2024
Approved by: Engineering Manager

SCOPE
This procedure covers routine maintenance and inspection of pump P-101.

SAFETY PRECAUTIONS
- Isolate all energy sources before beginning work
- Wear PPE at all times (gloves, safety glasses, hard hat)
- Verify zero-energy state before opening pump casing
- Ensure hot-work permit is obtained if welding is required

PROCEDURE
Step 1: Shut down pump P-101 using control panel HMI.
Step 2: Close suction valve SV-101 and discharge valve DV-101.
Step 3: Drain pump casing via drain valve.
Step 4: Inspect impeller for wear or damage.
Step 5: Check mechanical seal for leaks.
Step 6: Torque bolts to 80 Nm per updated specification (CHANGED from Rev 3).
Step 7: Reconnect and perform leak test at 80% operating pressure.
Step 8: Log maintenance record in CMMS system.

NOTE: Rev 4 updates bolt torque from 50 Nm to 80 Nm. Old specification is SUPERSEDED.

SIGN-OFF
Technician: _______________  Date: _______________
Supervisor: _______________  Date: _______________
""",
        """REVISION HISTORY
Rev 4 | 15/01/2024 | CRITICAL: Updated bolt torque to 80 Nm (was 50 Nm)
Rev 3 | 10/06/2023 | Updated safety section
Rev 2 | 15/03/2022 | Added drain step
Rev 1 | 01/01/2021 | Initial release

EQUIPMENT TAGS REFERENCED
P-101, SV-101, DV-101, HV-204, FCV-301
"""
    ])


def create_inspection_sheet_pdf():
    """Inspection checklist for heat exchanger E-201."""
    _new_pdf("inspection_checklist_e201.pdf", [
        """HEAT EXCHANGER INSPECTION CHECKLIST
Equipment: E-201 (Shell & Tube Heat Exchanger)
Date: 20/03/2024
Inspector: R. Sharma
Revision 2

VISUAL INSPECTION
------------------------------------------------------------
Item                      | Status   | Notes
Tube sheet condition      | PASS     | No corrosion found
Shell flanges             | PASS     | Gaskets in good condition
Support structures        | PASS     | Bolts tightened to spec
Instrumentation PT-201    | FAIL     | Calibration due
Insulation integrity      | PASS     | Minor damage at inlet

PRESSURE TEST
Test Pressure: 12 bar
Duration: 30 minutes
Result: PASS — No leaks detected

FOULING ASSESSMENT
Shell side: Low fouling (within acceptable limits)
Tube side: Moderate fouling — cleaning scheduled

Next inspection due: 20/09/2024

Equipment IDs: E-201, PT-201, HX-201
"""
    ])


def create_oem_manual_excerpt():
    """OEM manual excerpt for control valve FCV-301."""
    _new_pdf("oem_manual_fcv301_excerpt.pdf", [
        """OEM TECHNICAL MANUAL
Manufacturer: FlowControl Industries
Model: FCV-301 Globe Control Valve
Document No: FCI-MAN-2023-0847
Version 2.1
Date: 05/09/2023

PRODUCT DESCRIPTION
The FCV-301 is a globe-style control valve designed for precise flow regulation
in industrial process plants. Operating pressure: 0-25 bar. Temperature range:
-20°C to 250°C.

INSTALLATION REQUIREMENTS
1. Ensure pipe is flushed before installation.
2. Install valve with flow arrow in direction of process flow.
3. Provide 5x pipe diameters upstream and 3x downstream.
4. Torque body bolts to 65 Nm.

MAINTENANCE SCHEDULE
Every 6 months: Inspect packing gland, check for stem leakage.
Every 12 months: Full disassembly, inspect trim, replace packing.
Every 36 months: Replace actuator diaphragm.

TROUBLESHOOTING
Symptom: Valve won't close fully
Cause: Debris in trim
Action: Remove and clean trim assembly.

Equipment: FCV-301, FCV-302, PT-301
""",
        """SPARE PARTS LIST (FCV-301)
Part No.    Description               Qty  Lead Time
FCI-001     Valve body seal kit        1    2 weeks
FCI-002     Actuator diaphragm         1    4 weeks
FCI-003     Stem packing set           2    1 week
FCI-004     Trim assembly              1    6 weeks

CONTACT: techsupport@flowcontrol.example.com
Revision History: v2.1 (2023) — Updated torque specs
"""
    ])


def create_email_archive():
    """Simulated maintenance email thread."""
    _new_pdf("email_archive_p101_maintenance.pdf", [
        """EMAIL ARCHIVE — MAINTENANCE THREAD
Exported: 25/06/2024

========================================================
From: r.sharma@plant.example.com
To: maintenance@plant.example.com
CC: engineer@plant.example.com
Date: 12/06/2024 09:14
Subject: RE: P-101 Recurring Seal Failure

Team,

Following up on the P-101 issue. We've now had 3 seal failures in 6 months.
The mechanical seal on P-101 failed again last night around 22:30.

Root cause analysis suggests the seal material (VITON) may not be compatible
with the new process fluid introduced in March 2024 (pH < 3.5).

Recommendation: Switch to PTFE seal material. Cost: ~$450. Lead time: 2 weeks.

Action items:
- Raise work order for P-101 seal replacement using PTFE
- Review fluid compatibility across all pumps in the P-100 series
- Update SOP to specify PTFE seal material

Regards,
R. Sharma (Process Engineer)

========================================================
From: engineer@plant.example.com
To: maintenance@plant.example.com
Date: 12/06/2024 11:02
Subject: RE: P-101 Recurring Seal Failure

Approved. Please raise PO for PTFE seals.
Also check P-102 — same service, same risk.

Equipment referenced: P-101, P-102, HV-204
"""
    ])


def create_work_orders_csv():
    """Realistic work order CSV with mixed formatting."""
    content = """EqpNum,Fault,Date,Tech,Notes
P-101,Mechanical seal leak,12/06/2024,R. Sharma,Replaced VITON seal. Recurring issue - recommend PTFE upgrade.
HV-204,Actuator not responding,10/06/2024,J. Kumar,Replaced solenoid coil. Tested OK.
FCV-301,Valve hunting unstable,08/06/2024,A. Patel,Tuned PID controller. Reduced gain from 2.5 to 1.8.
 p-102 ,Vibration high,05/06/2024,R. Sharma,Realigned coupling. Vibration reduced from 12mm/s to 3mm/s.
E-201,Fouling on tube side,01/06/2024,B. Singh,Chemical cleaning performed. Heat duty restored to 98%.
,Pressure gauge faulty,28/05/2024,J. Kumar,Gauge replaced. Equipment tag unclear - requires manual review.
T-401,Level transmitter drift,25/05/2024,A. Patel,Recalibrated LT-401. Zero drift corrected.
"""
    out = OUTPUT_DIR / "work_orders_june2024.csv"
    out.write_text(content, encoding="utf-8")
    print(f"  Created: {out.name}")


def create_inspection_xlsx():
    """Inspection sheet as Excel with merged cells."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inspection Results"

        # Headers
        ws["A1"] = "MONTHLY EQUIPMENT INSPECTION REPORT"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")

        ws["A2"] = "Month: June 2024"
        ws["D2"] = "Plant: Unit 3"

        headers = ["Equipment ID", "Description", "Status", "Inspector", "Date", "Notes"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True)

        data = [
            ["P-101", "Centrifugal Pump", "FAIL", "R. Sharma", "01/06/2024", "Seal leaking"],
            ["P-102", "Centrifugal Pump", "PASS", "R. Sharma", "01/06/2024", "OK"],
            ["HV-204", "On/Off Valve", "PASS", "J. Kumar", "02/06/2024", "Actuator OK"],
            ["FCV-301", "Control Valve", "PASS", "A. Patel", "03/06/2024", "PID stable"],
            ["E-201", "Heat Exchanger", "WARN", "B. Singh", "04/06/2024", "Fouling - monitor"],
            ["T-401", "Storage Tank", "PASS", "A. Patel", "05/06/2024", "Level TX recalibrated"],
        ]

        for row_idx, row in enumerate(data, 4):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        out = OUTPUT_DIR / "monthly_inspection_june2024.xlsx"
        wb.save(str(out))
        print(f"  Created: {out.name}")
    except ImportError:
        print("  Skipped XLSX (openpyxl not installed)")


if __name__ == "__main__":
    print("Generating demo corpus...")
    create_sop_rev3()
    create_sop_rev4()
    create_inspection_sheet_pdf()
    create_oem_manual_excerpt()
    create_email_archive()
    create_work_orders_csv()
    create_inspection_xlsx()
    print("\nDone! Demo corpus created in demo_docs/")
    print("Files:")
    for f in sorted(OUTPUT_DIR.iterdir()):
        if f.suffix in (".pdf", ".csv", ".xlsx"):
            print(f"  {f.name} ({f.stat().st_size} bytes)")
